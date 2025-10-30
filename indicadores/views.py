from django.shortcuts import render
from django.views.generic import ListView, DetailView, CreateView, UpdateView
from django.urls import reverse_lazy
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.db.models import Q
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import datetime

from .models import Indicador, MedicaoIndicador
from .forms import IndicadorForm, MedicaoIndicadorForm

#class IndicadorListView(LoginRequiredMixin, ListView):
class IndicadorListView(ListView):
    model = Indicador
    template_name = 'indicadores/indicador_list.html'
    context_object_name = 'indicadores'
    paginate_by = 10
    
    def get_queryset(self):
        queryset = Indicador.objects.all().order_by('nome')
        
        # Filtros
        self.norma_filter = self.request.GET.get('norma', '')
        self.categoria_filter = self.request.GET.get('categoria', '')
        self.tipo_filter = self.request.GET.get('tipo', '')
        self.status_filter = self.request.GET.get('status', '')
        
        if self.norma_filter:
            queryset = queryset.filter(norma=self.norma_filter)
        if self.categoria_filter:
            queryset = queryset.filter(categoria=self.categoria_filter)
        if self.tipo_filter:
            queryset = queryset.filter(tipo=self.tipo_filter)
        if self.status_filter:
            if self.status_filter == 'acima_meta':
                # Filtra indicadores acima da meta
                indicadores_ids = []
                for indicador in queryset:
                    ultima_medicao = indicador.medicaoindicador_set.last()
                    if ultima_medicao and ultima_medicao.valor_medido >= indicador.meta:
                        indicadores_ids.append(indicador.id)
                queryset = queryset.filter(id__in=indicadores_ids)
            elif self.status_filter == 'abaixo_meta':
                # Filtra indicadores abaixo da meta
                indicadores_ids = []
                for indicador in queryset:
                    ultima_medicao = indicador.medicaoindicador_set.last()
                    if ultima_medicao and ultima_medicao.valor_medido < indicador.meta:
                        indicadores_ids.append(indicador.id)
                queryset = queryset.filter(id__in=indicadores_ids)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Adicionar estatísticas
        indicadores = self.get_queryset()
        context['total_indicadores'] = indicadores.count()
        
        # Calcular indicadores acima e abaixo da meta
        acima_meta = 0
        abaixo_meta = 0
        
        for indicador in indicadores:
            ultima_medicao = indicador.medicaoindicador_set.last()
            if ultima_medicao:
                if ultima_medicao.valor_medido >= indicador.meta:
                    acima_meta += 1
                else:
                    abaixo_meta += 1
        
        context['acima_meta'] = acima_meta
        context['abaixo_meta'] = abaixo_meta
        
        # Adicionar filtros atuais ao contexto
        context['norma_filter'] = self.norma_filter
        context['categoria_filter'] = self.categoria_filter
        context['tipo_filter'] = self.tipo_filter
        context['status_filter'] = self.status_filter
        
        return context

#class IndicadorDetailView(LoginRequiredMixin, DetailView):
class IndicadorDetailView(DetailView):
    model = Indicador
    template_name = 'indicadores/indicador_detail.html'
    context_object_name = 'indicador'

#class IndicadorCreateView(LoginRequiredMixin, CreateView):
class IndicadorCreateView(CreateView):
    model = Indicador
    form_class = IndicadorForm
    template_name = 'indicadores/indicador_form.html'
    success_url = reverse_lazy('indicadores:indicador-list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Indicador criado com sucesso!')
        return super().form_valid(form)

#class IndicadorUpdateView(LoginRequiredMixin, UpdateView):
class IndicadorUpdateView(UpdateView):
    model = Indicador
    form_class = IndicadorForm
    template_name = 'indicadores/indicador_form.html'
    
    def get_success_url(self):
        return reverse_lazy('indicadores:indicador-detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'Indicador atualizado com sucesso!')
        return super().form_valid(form)

# Views para exportação
def exportar_indicadores_pdf(request):
    # Obter parâmetros de filtro
    norma = request.GET.get('norma', '')
    categoria = request.GET.get('categoria', '')
    tipo = request.GET.get('tipo', '')
    status = request.GET.get('status', '')
    
    # Aplicar filtros
    indicadores = Indicador.objects.all()
    if norma:
        indicadores = indicadores.filter(norma=norma)
    if categoria:
        indicadores = indicadores.filter(categoria=categoria)
    if tipo:
        indicadores = indicadores.filter(tipo=tipo)
    
    # Criar o PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Centro
    )
    
    # Título
    title = Paragraph("Relatório de Indicadores - Zenith SGI", title_style)
    elements.append(title)
    
    # Informações do relatório
    info_text = f"Data de emissão: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}<br/>"
    info_text += f"Total de indicadores: {indicadores.count()}<br/>"
    if norma:
        info_text += f"Norma: {dict(Indicador.NORMA_CHOICES).get(norma, norma)}<br/>"
    if categoria:
        info_text += f"Categoria: {dict(Indicador.CATEGORIA_CHOICES).get(categoria, categoria)}<br/>"
    if tipo:
        info_text += f"Tipo: {dict(Indicador.TIPO_CHOICES).get(tipo, tipo)}<br/>"
    
    info_paragraph = Paragraph(info_text, styles['Normal'])
    elements.append(info_paragraph)
    elements.append(Spacer(1, 20))
    
    # Tabela de dados
    data = [['Nome', 'Tipo', 'Meta', 'Última Medição', 'Status', 'Responsável']]
    
    for indicador in indicadores:
        ultima_medicao = indicador.medicaoindicador_set.last()
        valor_ultima = ultima_medicao.valor_medido if ultima_medicao else '-'
        
        if ultima_medicao:
            if ultima_medicao.valor_medido >= indicador.meta:
                status_text = 'Acima da Meta'
            else:
                status_text = 'Abaixo da Meta'
        else:
            status_text = 'Sem Dados'
        
        data.append([
            indicador.nome,
            indicador.get_tipo_display(),
            f"{indicador.meta} {indicador.unidade_medida}",
            f"{valor_ultima} {indicador.unidade_medida}" if valor_ultima != '-' else '-',
            status_text,
            indicador.responsavel
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    
    # Gerar PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio_indicadores.pdf"'
    return response

def exportar_indicadores_excel(request):
    # Obter parâmetros de filtro
    norma = request.GET.get('norma', '')
    categoria = request.GET.get('categoria', '')
    tipo = request.GET.get('tipo', '')
    status = request.GET.get('status', '')
    
    # Aplicar filtros
    indicadores = Indicador.objects.all()
    if norma:
        indicadores = indicadores.filter(norma=norma)
    if categoria:
        indicadores = indicadores.filter(categoria=categoria)
    if tipo:
        indicadores = indicadores.filter(tipo=tipo)
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Indicadores"
    
    # Cabeçalho
    headers = ['Nome', 'Norma', 'Categoria', 'Tipo', 'Meta', 'Última Medição', 'Status', 'Unidade', 'Frequência', 'Responsável']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Dados
    for row, indicador in enumerate(indicadores, 2):
        ultima_medicao = indicador.medicaoindicador_set.last()
        valor_ultima = ultima_medicao.valor_medido if ultima_medicao else '-'
        
        if ultima_medicao:
            if ultima_medicao.valor_medido >= indicador.meta:
                status_text = 'Acima da Meta'
            else:
                status_text = 'Abaixo da Meta'
        else:
            status_text = 'Sem Dados'
        
        ws.cell(row=row, column=1, value=indicador.nome)
        ws.cell(row=row, column=2, value=indicador.get_norma_display())
        ws.cell(row=row, column=3, value=indicador.get_categoria_display())
        ws.cell(row=row, column=4, value=indicador.get_tipo_display())
        ws.cell(row=row, column=5, value=float(indicador.meta))
        ws.cell(row=row, column=6, value=float(valor_ultima) if valor_ultima != '-' else '-')
        ws.cell(row=row, column=7, value=status_text)
        ws.cell(row=row, column=8, value=indicador.unidade_medida)
        ws.cell(row=row, column=9, value=indicador.frequencia_medicao)
        ws.cell(row=row, column=10, value=indicador.responsavel)
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Preparar resposta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="relatorio_indicadores.xlsx"'
    
    wb.save(response)
    return response