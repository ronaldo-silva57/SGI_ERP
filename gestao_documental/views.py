from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.http import JsonResponse
from django.db.models import Count
from .models import Documento, CategoriaDocumento
from .forms import DocumentoForm, CategoriaDocumentoForm
import io #pdf xls
from django.db.models import Q#pdf xls
from django.http import HttpResponse #pdf xls ok
from openpyxl import Workbook #xls
from openpyxl.styles import Font, Alignment, PatternFill #xls
import datetime#pdf xls
from reportlab.lib import colors#pdf xls
from reportlab.lib.pagesizes import A4#pdf xls
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer#pdf xls
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle#pdf xls

class DocumentoListView(ListView):
    model = Documento
    template_name = 'gestao_documental/documento_list.html'
    context_object_name = 'documentos'
    paginate_by = 10
    
    def get_queryset(self):
        queryset = Documento.objects.all()
        norma = self.request.GET.get('norma')
        status = self.request.GET.get('status')
        categoria = self.request.GET.get('categoria')
        tipo_documento = self.request.GET.get('tipo_documento')
        
        if norma:
            queryset = queryset.filter(norma=norma)
        if status:
            queryset = queryset.filter(status=status)
        if categoria:
            queryset = queryset.filter(categoria_id=categoria)
        if tipo_documento:
            queryset = queryset.filter(tipo_documento=tipo_documento)
            
        return queryset.order_by('-data_elaboracao')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categorias'] = CategoriaDocumento.objects.all()
        return context

class DocumentoDetailView(DetailView):
    model = Documento
    template_name = 'gestao_documental/documento_detail.html'
    context_object_name = 'documento'

class DocumentoCreateView(CreateView):
    model = Documento
    form_class = DocumentoForm
    template_name = 'gestao_documental/documento_form.html'
    success_url = reverse_lazy('gestao_documental:documento-list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Documento criado com sucesso!')
        return super().form_valid(form)

class DocumentoUpdateView(UpdateView):
    model = Documento
    form_class = DocumentoForm
    template_name = 'gestao_documental/documento_form.html'
    
    def get_success_url(self):
        return reverse_lazy('gestao_documental:documento-detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'Documento atualizado com sucesso!')
        return super().form_valid(form)

# Views para CategoriaDocumento
class CategoriaDocumentoListView(ListView):
    model = CategoriaDocumento
    template_name = 'gestao_documental/categoria_list.html'
    context_object_name = 'categorias'
    paginate_by = 10

class CategoriaDocumentoCreateView(CreateView):
    model = CategoriaDocumento
    form_class = CategoriaDocumentoForm
    template_name = 'gestao_documental/categoria_form.html'
    success_url = reverse_lazy('gestao_documental:categoria-list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Categoria criada com sucesso!')
        return super().form_valid(form)

class CategoriaDocumentoUpdateView(UpdateView):
    model = CategoriaDocumento
    form_class = CategoriaDocumentoForm
    template_name = 'gestao_documental/categoria_form.html'
    success_url = reverse_lazy('gestao_documental:categoria-list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Categoria atualizada com sucesso!')
        return super().form_valid(form)

class CategoriaDocumentoDeleteView(DeleteView):
    model = CategoriaDocumento
    template_name = 'gestao_documental/categoria_confirm_delete.html'
    success_url = reverse_lazy('gestao_documental:categoria-list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Categoria excluída com sucesso!')
        return super().delete(request, *args, **kwargs)

@login_required
def documento_dashboard(request):
    total_documentos = Documento.objects.count()
    documentos_aprovados = Documento.objects.filter(status='aprovado').count()
    documentos_revisao = Documento.objects.filter(status='revisao').count()
    
    # Estatísticas por norma
    normas_stats = {}
    for norma_code, norma_name in Documento.NORMA_CHOICES:
        count = Documento.objects.filter(norma=norma_code).count()
        normas_stats[norma_name] = count
    
    # Estatísticas por categoria
    categorias_stats = CategoriaDocumento.objects.annotate(
        total_documentos=Count('documento')
    ).values('nome', 'total_documentos')
    
    # Estatísticas por tipo de documento
    tipos_stats = {}
    for tipo_code, tipo_name in Documento.TIPO_CHOICES:
        count = Documento.objects.filter(tipo_documento=tipo_code).count()
        tipos_stats[tipo_name] = count

    context = {
        'total_documentos': total_documentos,
        'documentos_aprovados': documentos_aprovados,
        'documentos_revisao': documentos_revisao,
        'normas_stats': normas_stats,
        'categorias_stats': categorias_stats,
        'tipos_stats': tipos_stats,
    }
    return render(request, 'gestao_documental/documento_dashboard.html', context)

def get_categoria_documentos(request, categoria_id):
    """Retorna documentos por categoria para AJAX"""
    documentos = Documento.objects.filter(categoria_id=categoria_id).values('id', 'codigo', 'titulo')
    return JsonResponse(list(documentos), safe=False)


#====== exportar pdf =======
def exportar_documentos_pdf(request):
    # Obter parâmetros de filtro da URL
    norma = request.GET.get('norma', '')
    status = request.GET.get('status', '')
    categoria_id = request.GET.get('categoria', '')
    tipo_documento = request.GET.get('tipo_documento', '')

    # Aplicar filtros
    documentos = Documento.objects.all()
    if norma:
        documentos = documentos.filter(norma=norma)
    if status:
        documentos = documentos.filter(status=status)
    if categoria_id:
        documentos = documentos.filter(categoria_id=categoria_id)
    if tipo_documento:
        documentos = documentos.filter(tipo_documento=tipo_documento)

    # --- Configuração do PDF ---
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []

    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Centro
    )
    
    # Título
    title = Paragraph("Relatório de Documentos - Gestão Documental", title_style)
    elements.append(title)

    # Informações do Relatório
    info_text = f"Data de emissão: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}<br/>"
    info_text += f"Total de documentos: {documentos.count()}<br/>"
    
    # Adicionar filtros ao rodapé (requer acesso às choices do modelo)
    norma_dict = dict(Documento.NORMA_CHOICES)
    status_dict = dict(Documento.STATUS_CHOICES)
    tipo_dict = dict(Documento.TIPO_CHOICES)
    
    if norma:
        info_text += f"Norma: {norma_dict.get(norma, norma)}<br/>"
    if status:
        info_text += f"Status: {status_dict.get(status, status)}<br/>"
    if categoria_id:
        try:
            categoria = CategoriaDocumento.objects.get(pk=categoria_id).nome
            info_text += f"Categoria: {categoria}<br/>"
        except CategoriaDocumento.DoesNotExist:
            pass # Ignora se a categoria não existir
    if tipo_documento:
        info_text += f"Tipo: {tipo_dict.get(tipo_documento, tipo_documento)}<br/>"
    
    info_paragraph = Paragraph(info_text, styles['Normal'])
    elements.append(info_paragraph)
    elements.append(Spacer(1, 20))

    # Tabela de dados
    data = [
        ['Código', 'Título', 'Tipo', 'Revisão', 'Status', 'Elaborador', 'Data Aprovação']
    ]

    for documento in documentos:
        data.append([
            documento.codigo,
            documento.titulo,
            documento.get_tipo_documento_display(),
            str(documento.revisao),
            documento.get_status_display(),
            documento.elaborador.get_full_name() or documento.elaborador.username,
            documento.data_aprovacao.strftime('%d/%m/%Y') if documento.data_aprovacao else '-'
        ])

    # Configuração de Estilo da Tabela
    # Ajustar largura da tabela para usar o máximo possível da página
    table = Table(data, colWidths=[60, 160, 60, 50, 80, 80, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (3, 0), (4, -1), 'CENTER'), # Alinhar Revisão e Status ao centro
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 1), (-1, -1), 9), # Tamanho da fonte para dados
    ]))

    elements.append(table)

    # Gerar PDF
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio_documentos.pdf"'
    return response

#======= exportar xls =======
def exportar_documentos_excel(request):
    # Obter parâmetros de filtro da URL
    norma = request.GET.get('norma', '')
    status = request.GET.get('status', '')
    categoria_id = request.GET.get('categoria', '')
    tipo_documento = request.GET.get('tipo_documento', '')

    # Aplicar filtros
    documentos = Documento.objects.all()
    if norma:
        documentos = documentos.filter(norma=norma)
    if status:
        documentos = documentos.filter(status=status)
    if categoria_id:
        documentos = documentos.filter(categoria_id=categoria_id)
    if tipo_documento:
        documentos = documentos.filter(tipo_documento=tipo_documento)

    # --- Configuração do Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Documentos"

    # Cabeçalho
    headers = [
        'Código', 'Título', 'Norma', 'Categoria', 'Tipo', 'Revisão', 
        'Status', 'Data Elaboração', 'Data Aprovação', 'Elaborador', 'Aprovador'
    ]
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Dados
    for row, documento in enumerate(documentos, 2):
        # Campos que podem ser nulos
        aprovador_nome = documento.aprovador.get_full_name() or documento.aprovador.username if documento.aprovador else '-'
        data_aprovacao = documento.data_aprovacao.strftime('%d/%m/%Y') if documento.data_aprovacao else '-'

        data_row = [
            documento.codigo,
            documento.titulo,
            documento.get_norma_display(),
            documento.categoria.nome, # Assumindo que categoria tem um campo 'nome'
            documento.get_tipo_documento_display(),
            documento.revisao,
            documento.get_status_display(),
            documento.data_elaboracao.strftime('%d/%m/%Y'),
            data_aprovacao,
            documento.elaborador.get_full_name() or documento.elaborador.username,
            aprovador_nome,
        ]
        
        for col, value in enumerate(data_row, 1):
            ws.cell(row=row, column=col, value=value)

    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                # O 'titulo' pode ser longo, ajusta o limite para um tamanho razoável
                if column_letter == 'B': # Coluna Título
                    max_length = max(max_length, min(45, len(str(cell.value))))
                else:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Preparar resposta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="relatorio_documentos.xlsx"'
    
    wb.save(response)
    return response